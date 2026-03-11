import streamlit as st
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
import io

# Проверка наличия openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Настройка страницы
st.set_page_config(
    page_title="Кредитный калькулятор",
    page_icon="💰",
    layout="wide"  # Широкий режим для лучшего отображения
)
st.title("💰 Кредитный калькулятор")

# Боковая панель с параметрами
with st.sidebar:
    st.header("Параметры кредита")

    amount = st.number_input("Сумма кредита (₽)", min_value=1000, value=1000000, step=10000)
    rate = st.number_input("Ставка (% годовых)", min_value=0.1, max_value=30.0, value=10.0, step=0.1)

    col1, col2 = st.columns(2)
    with col1:
        years = st.number_input("Лет", min_value=0, value=1, step=1)
    with col2:
        months = st.number_input("Месяцев", min_value=0, max_value=11, value=0, step=1)

    total_months = years * 12 + months
    if total_months == 0:
        st.error("Срок должен быть больше 0")
        st.stop()

    payment_type = st.radio("Тип платежа", ["Аннуитетный", "Дифференцированный"])

    use_dates = st.checkbox("Добавить даты платежей")
    if use_dates:
        start_date = st.date_input("Дата первого платежа", datetime.now().date() + relativedelta(months=1))

# Расчеты
monthly_rate = rate / 100 / 12


def calculate_schedule(amount, rate, months, type, start_date=None):
    schedule = []
    debt = amount

    if type == "Аннуитетный":
        if rate == 0:
            payment = amount / months
        else:
            payment = amount * (rate * (1 + rate) ** months) / ((1 + rate) ** months - 1)

        for month in range(1, months + 1):
            interest = debt * rate
            principal = payment - interest

            if principal > debt:
                principal = debt
                payment = debt + interest

            debt_after = debt - principal
            row = {
                'Месяц': month,
                'Остаток на начало': round(debt, 2),
                'Платеж': round(payment, 2),
                'Проценты': round(interest, 2),
                'Основной долг': round(principal, 2),
                'Остаток на конец': round(max(debt_after, 0), 2)
            }
            if start_date:
                row['Дата'] = (start_date + relativedelta(months=month - 1)).strftime('%d.%m.%Y')

            schedule.append(row)
            debt = debt_after

    else:  # Дифференцированный
        principal_part = amount / months

        for month in range(1, months + 1):
            interest = debt * rate
            payment = principal_part + interest

            if month == months:
                payment = debt + interest
                principal_part = debt
                debt_after = 0
            else:
                debt_after = debt - principal_part

            row = {
                'Месяц': month,
                'Остаток на начало': round(debt, 2),
                'Платеж': round(payment, 2),
                'Проценты': round(interest, 2),
                'Основной долг': round(principal_part, 2),
                'Остаток на конец': round(max(debt_after, 0), 2)
            }
            if start_date:
                row['Дата'] = (start_date + relativedelta(months=month - 1)).strftime('%d.%m.%Y')

            schedule.append(row)
            debt = debt_after

    return pd.DataFrame(schedule)


# Создаем график платежей
df = calculate_schedule(amount, monthly_rate, total_months, payment_type,
                        start_date if use_dates else None)

# Отображаем результаты
st.header(" Результаты")

total_payment = df['Платеж'].sum()
total_interest = df['Проценты'].sum()

# ИСПРАВЛЕНО: увеличенные метрики с подсказками
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric(
        "Общая сумма выплат",
        f"{total_payment:,.0f} ₽",
        help="Полная сумма, которую вы вернете банку"
    )
with col2:
    overpayment_percent = total_interest / amount * 100
    st.metric(
        "Переплата",
        f"{total_interest:,.0f} ₽",
        delta=f"{overpayment_percent:.1f}%",
        delta_color="inverse",
        help="Сумма процентов + разница в платежах"
    )
with col3:
    if payment_type == "Аннуитетный":
        monthly_payment = df['Платеж'].iloc[0]
        st.metric(
            "Ежемесячный платеж",
            f"{monthly_payment:,.0f} ₽",
            help="Фиксированный платеж каждый месяц"
        )
    else:
        first_payment = df['Платеж'].iloc[0]
        last_payment = df['Платеж'].iloc[-1]
        st.metric(
            "Платеж: первый/последний",
            f"{first_payment:,.0f} / {last_payment:,.0f} ₽",
            help="Первый и последний месячные платежи"
        )
with col4:
    st.metric(
        "Срок кредита",
        f"{total_months} мес.",
        help=f"{years} лет {months} месяцев"
    )

# График
st.subheader(" График платежей")
chart_df = df[['Месяц', 'Основной долг', 'Проценты']].set_index('Месяц')
st.area_chart(chart_df)


st.subheader(" Детальный график")

# Создаем копию для отображения с форматированием
display_df = df.copy()

# Форматируем числовые колонки
for col in display_df.columns:
    if col not in ['Месяц', 'Дата']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.2f} ₽")

# Настраиваем конфигурацию колонок для лучшего отображения
if use_dates:
    column_config = {
        "Месяц": st.column_config.NumberColumn("№", width="small"),
        "Дата": st.column_config.TextColumn("Дата платежа", width="small"),
        "Остаток на начало": st.column_config.TextColumn("Остаток на начало", width="medium"),
        "Платеж": st.column_config.TextColumn("Ежемесячный платеж", width="medium"),
        "Проценты": st.column_config.TextColumn("Проценты", width="medium"),
        "Основной долг": st.column_config.TextColumn("Основной долг", width="medium"),
        "Остаток на конец": st.column_config.TextColumn("Остаток на конец", width="medium")
    }
else:
    column_config = {
        "Месяц": st.column_config.NumberColumn("№", width="small"),
        "Остаток на начало": st.column_config.TextColumn("Остаток на начало", width="medium"),
        "Платеж": st.column_config.TextColumn("Ежемесячный платеж", width="medium"),
        "Проценты": st.column_config.TextColumn("Проценты", width="medium"),
        "Основной долг": st.column_config.TextColumn("Основной долг", width="medium"),
        "Остаток на конец": st.column_config.TextColumn("Остаток на конец", width="medium")
    }

# Отображаем таблицу с настройками
st.dataframe(
    display_df,
    column_config=column_config,
    use_container_width=True,
    hide_index=True,
    height=400  # Фиксированная высота с прокруткой
)

# Кнопки для скачивания
st.subheader("💾 Скачать график")

col1, col2 = st.columns(2)

with col1:
    # CSV
    csv = df.to_csv(index=False, encoding='utf-8-sig')
    st.download_button(
        " Скачать CSV",
        csv,
        f"credit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        use_container_width=True
    )

with col2:
    # Excel
    if OPENPYXL_AVAILABLE:
        # Создаем Excel файл
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='График платежей', index=False)

            # Форматирование
            workbook = writer.book
            worksheet = writer.sheets['График платежей']

            # Заголовки
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Ширина колонок (увеличена для сумм)
            for col in worksheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # Увеличиваем ширину для денежных колонок
                worksheet.column_dimensions[col_letter].width = min(max_length + 5, 40)

        output.seek(0)

        st.download_button(
            " Скачать Excel",
            output,
            f"credit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.info("Для Excel установите: pip install openpyxl")
        if st.button("📦 Показать команду"):
            st.code("pip install openpyxl", language="bash")