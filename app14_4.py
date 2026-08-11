
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

selected_file_path = None

def process_excel(file_path):
    """Суммирует числа в первом столбце (пропуская заголовок)"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    total = 0
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, (int, float)):
            total += cell.value
    return total

def process_search_copy(file_path):
    """Ищет 'проба' в столбце A и копирует строки на лист 'тест'"""
    wb = openpyxl.load_workbook(file_path)

    # Определяем исходный лист
    if 'Лист1' in wb.sheetnames:
        ws_source = wb['Лист1']
    else:
        ws_source = wb.active  # если "Лист1" нет, берём активный

    # Удаляем старый лист "тест", если он есть
    if 'тест' in wb.sheetnames:
        del wb['тест']

    ws_target = wb.create_sheet('тест')
    copied = 0

    # Проходим по всем строкам исходного листа
    for row in ws_source.iter_rows(min_row=1, max_col=ws_source.max_column):
        cell_a = row[0]  # первый столбец
        if cell_a.value and isinstance(cell_a.value, str) and 'проба' in cell_a.value.lower():
            # Копируем значения всей строки
            values = [c.value for c in row]
            ws_target.append(values)
            copied += 1

    wb.save(file_path)
    return copied

def select_file():
    global selected_file_path
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        selected_file_path = file_path
        label_file.config(text=f"Выбран: {file_path.split('/')[-1]}")

def sum_column():
    if not selected_file_path:
        messagebox.showwarning("Файл не выбран", "Сначала выберите файл.")
        return
    try:
        result = process_excel(selected_file_path)
        messagebox.showinfo("Результат", f"Сумма чисел в первом столбце: {result}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")

def search_proba():
    if not selected_file_path:
        messagebox.showwarning("Файл не выбран", "Сначала выберите файл.")
        return
    try:
        copied = process_search_copy(selected_file_path)
        messagebox.showinfo("Готово", f"Создан лист 'тест', скопировано строк: {copied}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")

# Интерфейс
root = tk.Tk()
root.title("Обработка Excel")
root.geometry("400x250")

label_file = tk.Label(root, text="Файл не выбран", fg="gray")
label_file.pack(pady=10)

btn_select = tk.Button(root, text="Выбрать файл", command=select_file)
btn_select.pack(pady=5)

btn_sum = tk.Button(root, text="Сумма по столбцу A", command=sum_column)
btn_sum.pack(pady=5)

btn_search = tk.Button(root, text="Поиск 'проба' и копирование в лист 'тест'", command=search_proba)
btn_search.pack(pady=5)

root.mainloop()






import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    total = 0
    for row in ws.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, (int, float)):
            total += cell.value
    return total

def select_file():
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        try:
            result = process_excel(file_path)
            messagebox.showinfo("Результат", f"Сумма чисел в первом столбце: {result}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")

root = tk.Tk()
root.title("Обработка Excel")
root.geometry("300x150")

tk.Label(root, text="Выберите .xlsx файл для обработки").pack(pady=20)
tk.Button(root, text="Выбрать файл", command=select_file).pack(pady=10)

root.mainloop()





import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

def process_excel(file_path):
    """Обработка Excel-файла и возврат результата"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    total = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            total += row[0]
    return total

def select_file():
    """Диалог выбора файла и запуск обработки"""
    file_path = filedialog.askopenfilename(
        title="Выберите файл Excel",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if file_path:
        try:
            result = process_excel(file_path)
            messagebox.showinfo("Результат", f"Сумма чисел в первом столбце: {result}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")

# Создаём простое окно
root = tk.Tk()
root.title("Обработка Excel")
root.geometry("300x150")

label = tk.Label(root, text="Выберите .xlsx файл для обработки")
label.pack(pady=20)

button = tk.Button(root, text="Выбрать файл", command=select_file)
button.pack(pady=10)

root.mainloop()



Да, это абсолютно реально. Можно создать веб-приложение на Flask, которое при нажатии кнопки открывает стандартное диалоговое окно выбора файла (через <input type="file">), загружает .xlsx на сервер, обрабатывает его и показывает результат. Либо можно сделать настольное приложение на Python (например, на tkinter, PyQt), где диалоговое окно выбора файла также легко реализуется.

Я покажу веб-вариант на Flask — он продолжит ваш проект с чат-ботом, и вы сможете добавить к нему функциональность работы с Excel.

Как это будет работать

1. Пользователь заходит на страницу с формой загрузки.
2. Нажимает «Выбрать файл» → открывается системный диалог выбора .xlsx.
3. Файл отправляется на сервер, где обрабатывается с помощью библиотеки openpyxl (или pandas).
4. Можно, например, прочитать данные, изменить их, построить отчёт и показать результат на новой странице или вернуть изменённый файл для скачивания.

Что понадобится

· Установить библиотеку для работы с Excel:
  ```bash
  pip install openpyxl
  ```
  или, если нужна более мощная аналитика:
  ```bash
  pip install pandas openpyxl
  ```
· Стандартные Flask и шаблоны.

Пример реализации

1. Маршруты в app.py

```python
import os
from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/excel', methods=['GET', 'POST'])
def excel_tool():
    if request.method == 'POST':
        # Проверяем, что файл пришёл
        if 'file' not in request.files:
            return 'Файл не выбран'
        file = request.files['file']
        if file.filename == '':
            return 'Файл не выбран'
        if file and allowed_file(file.filename):
            # Сохраняем или сразу читаем в память
            wb = openpyxl.load_workbook(file)
            # --- ДЕЛАЕМ МАНИПУЛЯЦИИ ---
            # Например, считаем сумму в первом столбце
            ws = wb.active
            total = 0
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], (int, float)):
                    total += row[0]
            # Возвращаем результат на страницу
            return render_template('excel_result.html', total=total, filename=file.filename)
        else:
            return 'Разрешены только файлы .xlsx'
    return render_template('excel_upload.html')
```

2. Шаблон templates/excel_upload.html

```html
<!DOCTYPE html>
<html>
<head>
    <title>Загрузка Excel</title>
</head>
<body>
    <h1>Выберите .xlsx файл</h1>
    <form method="post" enctype="multipart/form-data">
        <!-- Это поле откроет диалог выбора файла -->
        <input type="file" name="file" accept=".xlsx" required>
        <button type="submit">Загрузить и обработать</button>
    </form>
</body>
</html>
```

3. Шаблон templates/excel_result.html

```html
<!DOCTYPE html>
<html>
<head>
    <title>Результат обработки</title>
</head>
<body>
    <h1>Файл {{ filename }} обработан</h1>
    <p>Сумма чисел в первом столбце: {{ total }}</p>
    <a href="/excel">Загрузить другой файл</a>
</body>
</html>
```

Какие манипуляции можно делать

· Читать и анализировать данные (сортировка, фильтрация, сводки).
· Изменять ячейки и сохранять новый файл (можно отдать на скачивание через send_file).
· Строить графики (с помощью matplotlib + openpyxl).
· Объединять несколько файлов, конвертировать форматы и т.п.

Важные моменты

· Безопасность: всегда проверяйте расширение файла и используйте secure_filename() для сохранения на диск, чтобы избежать path traversal.
· Большие файлы: если файлы очень большие, лучше читать их в потоке и не хранить целиком в памяти.
· Интерфейс: стандартное поле <input type="file"> во всех браузерах открывает системный диалог выбора файла.

Если хотите десктопное приложение с диалоговым окном, можно использовать tkinter.filedialog.askopenfilename(), но тогда нужен будет другой фреймворк. Учитывая ваш опыт с Flask, веб-вариант будет проще.

Таким образом, да, такое приложение можно сделать, и это довольно просто.
